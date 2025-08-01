from email import message
from genericpath import samefile
from django.contrib import messages
from django.http.response import HttpResponseRedirect
from django.shortcuts import render
from django.http import FileResponse, HttpResponse, JsonResponse
import pandas as pd

from store.views import numOfDays
from transactions.filters import inward_filter, outward_filter, stock_filter, supply_return_filter
from .forms import *
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from .models import *
from datetime import date
from django.core.paginator import Paginator, EmptyPage
from functools import reduce

from django.urls import reverse
import csv
import mimetypes

import os

# Build paths inside the project like this: os.path.join(BASE_DIR, ...)
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


from datetime import date
import pytz

IST = pytz.timezone('Asia/Kolkata')




def demo(request):

    
    
    s = stock.objects.all()

    inward_fi = inward.objects.all()
    outward_fi = outward.objects.all()
    supply_return_fi = supply_return.objects.all()


    for ab in s:

        a = inward_fi.filter(company__company_name = ab.company.company_name, company_goods__name = ab.company_goods.name, goods_company__goods_company_name = ab.goods_company.goods_company_name, DC_date__range=["2023-04-01", "2023-06-15"])
        b = outward_fi.filter(company__company_name =  ab.company.company_name, company_goods__name = ab.company_goods.name, goods_company__goods_company_name = ab.goods_company.goods_company_name, DC_date__range=["2023-04-01", "2023-06-15"])
        c = supply_return_fi.filter(company__company_name =  ab.company.company_name, company_goods__name = ab.company_goods.name, goods_company__goods_company_name = ab.goods_company.goods_company_name, DC_date__range=["2023-04-01", "2023-06-15"])

        x = 0
        y = 0
        z = 0

        for i in a:
            x = x + i.bags
            
        for i in b:
            y = y + i.bags

        for i in c:
            z = z + i.bags

        

        st = x - y + z


        ab.total_bag = st
        ab.save()

def delete_entry(request):

    inward.objects.filter(DC_date__lte = date(int(2023), int(3), int(31))).delete()
    outward.objects.filter(DC_date__lte = date(int(2023), int(3), int(31))).delete()
    supply_return.objects.filter(DC_date__lte = date(int(2023), int(3), int(31))).delete()





# Create your views here.

@login_required(login_url='login')
def add_inward(request):


    if request.method == 'POST':

        forms = inward_Form(request.POST)

        print(request.POST)

        if forms.is_valid():

          
            forms.save()

            return JsonResponse({'status' : 'done'}, safe=False)


        else:
                
            error = forms.errors.as_json()
            return JsonResponse({'error' : error}, safe=False)

    else:

        forms = inward_Form()

        context = {
            'form': forms
        }
        return render(request, 'transactions/add_inward.html', context)


@login_required(login_url='login')
def update_inward(request, inward_id ):


    if request.method == 'POST':

        instance_inward = inward.objects.get(id = inward_id)


        forms = inward_Form(request.POST, instance=instance_inward)


        if forms.is_valid():

            forms.save()

            return HttpResponseRedirect(reverse('list_inward'))

        else:

            instance = inward.objects.get(id = inward_id)
            comapnyID = forms.instance.company.id
            comapny_goods_ID = forms.instance.company_goods.id
            goods_company_ID = forms.instance.goods_company.id
            agent_ID = forms.instance.agent.id

            context = {
                'form': forms,
                'comapnyID' : comapnyID,
                'comapny_goods_ID' : comapny_goods_ID,
                'goods_company_ID' : goods_company_ID,
                'agent_ID' : agent_ID
            }
            

            return render(request, 'transactions/update_inward.html', context)


    else:

        instance = inward.objects.get(id = inward_id)
        forms = inward_Form(instance = instance)
        comapnyID = forms.instance.company.id
        comapny_goods_ID = forms.instance.company_goods.id
        goods_company_ID = forms.instance.goods_company.id
        agent_ID = forms.instance.agent.id

        context = {
            'form': forms,
            'comapnyID' : comapnyID,
            'comapny_goods_ID' : comapny_goods_ID,
            'goods_company_ID' : goods_company_ID,
            'agent_ID' : agent_ID
        }
        return render(request, 'transactions/update_inward.html', context)


@login_required(login_url='login')
def delete_inward(request, inward_id):

    try:
        inward.objects.filter(id = inward_id).delete()

        return HttpResponseRedirect(reverse('list_inward_delete'))


    except:
        return HttpResponseRedirect(reverse('list_inward_delete'))




@login_required(login_url='login')
def list_inward(request):

    year = request.GET.get('year')

    if year:

        date1 = year + '-04-01'
        date2 = str(int(year) + 1) + '-03-31'

        data = inward.objects.filter(DC_date__range=[date1, date2]).order_by("DC_number")
    
    else:

        data = inward.objects.filter().order_by("DC_number")

    total_bags = data.aggregate(Sum('bags'))['bags__sum']
    

    inward_filter_data = inward_filter()

    company_data = company.objects.all()

    page = request.GET.get('page', 1)
    paginator = Paginator(data, 50)

    try:
        data = paginator.page(page)
    except PageNotAnInteger:
        data = paginator.page(1)
    except EmptyPage:
        data = paginator.page(paginator.num_pages)

    context = {
        'data': data,
        'company_data' : company_data,
        'total_bags' : total_bags,
        'filter_inward' : inward_filter_data,
        'year' : year
    }

    return render(request, 'transactions/list_inward.html', context)

import json

@login_required(login_url='login')
def add_outward(request):

    if request.method == 'POST':

        forms = outward_Form(request.POST)
        DC_date = request.POST.get('DC_date')

      
        forms = outward_Form(request.POST)

        if forms.is_valid():

            forms.save()

            return JsonResponse({'status' : 'done'}, safe=False)



        else:

            error = forms.errors.as_json()
            return JsonResponse({'error' : error}, safe=False)


    else:

        forms = outward_Form()

        context = {
            'form': forms
        }
        return render(request, 'transactions/add_outward.html', context)


@login_required(login_url='login')
def report_dashbord(request):

    inward_filter_data = inward_filter()
    outward_filter_data = outward_filter()



    context = {
            'filter_inward': inward_filter_data,
            'filter_outward': outward_filter_data,
        }

    return render(request, 'transactions/report_dashbord.html', context)



from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

@login_required(login_url='login')
def list_outward(request):


    year = request.GET.get('year')

    if year:

        date1 =year + '-04-01'
        date2 = str(int(year) + 1) + '-03-31'

        data = outward.objects.filter(DC_date__range=[date1, date2]).order_by("DC_number")
    
    else:

        data = outward.objects.filter().order_by('DC_number')


    agent_name = request.GET.get('agent_name')

    if agent_name:

        data = data.filter(agent__name__icontains=agent_name)

    total_bags = data.aggregate(Sum('bags'))['bags__sum']
    

    outward_filter_data = outward_filter()

    company_data = company.objects.all()



    page = request.GET.get('page', 1)
    paginator = Paginator(data, 50)

    try:
        data = paginator.page(page)
    except PageNotAnInteger:
        data = paginator.page(1)
    except EmptyPage:
        data = paginator.page(paginator.num_pages)



    context = {
        'data': data,
        'filter_outward' : outward_filter_data,
        'total_bags' : total_bags,
        'company_data' : company_data,
        'year' : year

    }

    return render(request, 'transactions/list_outward.html', context)

@login_required(login_url='login')
def update_outward(request, outward_id):


    if request.method == 'POST':

        instance = outward.objects.get(id = outward_id)
      
       
        forms = outward_Form(request.POST, instance=instance)


        if forms.is_valid():

           
            forms.save()

            return HttpResponseRedirect(reverse('list_outward'))

        else:
           
            comapnyID = forms.instance.company.id
            comapny_goods_ID = forms.instance.company_goods.id
            goods_company_ID = forms.instance.goods_company.id
            agent_ID = forms.instance.agent.id

                
            context = {
                'form':  forms,
                'comapnyID' : comapnyID,
                'comapny_goods_ID' : comapny_goods_ID,
                'goods_company_ID' : goods_company_ID,
                'agent_ID' : agent_ID
            }

            return render(request, 'transactions/update_outward.html', context)


    else:

        instance = outward.objects.get(id = outward_id)
        forms = outward_Form(instance = instance)
        comapnyID = forms.instance.company.id
        comapny_goods_ID = forms.instance.company_goods.id
        goods_company_ID = forms.instance.goods_company.id
        agent_ID = forms.instance.agent.id

        context = {
            'form': forms,
            'comapnyID' : comapnyID,
            'comapny_goods_ID' : comapny_goods_ID,
            'goods_company_ID' : goods_company_ID,
            'agent_ID' : agent_ID
        }
        return render(request, 'transactions/update_outward.html', context)


@login_required(login_url='login')
def delete_outward(request, outward_id):

    try:
        outward.objects.get(id = outward_id).delete()

        return HttpResponseRedirect(reverse('list_outward_delete'))


    except:
        return HttpResponseRedirect(reverse('list_outward_delete'))




from django.db.models import Sum, F, ExpressionWrapper, IntegerField
from django.http import JsonResponse
import pandas as pd
from itertools import chain


@login_required(login_url='login')
def list_stock(request):

    # Get selected year from GET or session
    selected_year = request.GET.get('year')
    if selected_year:
        request.session['selected_year'] = selected_year
    else:
        selected_year = request.session.get('selected_year', datetime.now().year)

    selected_year = int(selected_year)

    # Set date range: 1st April selected_year to 31st March (selected_year + 1)
    from_date = datetime(selected_year, 4, 1)
    to_date = datetime(selected_year + 1, 3, 31, 23, 59, 59)

    inward_combinations = inward.objects.values('company', 'company_goods', 'goods_company').distinct()
    outward_combinations = outward.objects.values('company', 'company_goods', 'goods_company').distinct()
    supply_return_combinations = supply_return.objects.values('company', 'company_goods', 'goods_company').distinct()

    all_combinations = list(chain(inward_combinations, outward_combinations, supply_return_combinations))

    stock_dict = {}

    for combination in all_combinations:
        key = (combination['company'], combination['company_goods'], combination['goods_company'])

        inward_total = inward.objects.filter(
            company=combination['company'],
            company_goods=combination['company_goods'],
            goods_company=combination['goods_company'],
            DC_date__range=(from_date, to_date)
        ).aggregate(total=Sum('bags'))['total'] or 0

        outward_total = outward.objects.filter(
            company=combination['company'],
            company_goods=combination['company_goods'],
            goods_company=combination['goods_company'],
            DC_date__range=(from_date, to_date)
        ).aggregate(total=Sum('bags'))['total'] or 0

        return_total = supply_return.objects.filter(
            company=combination['company'],
            company_goods=combination['company_goods'],
            goods_company=combination['goods_company'],
            DC_date__range=(from_date, to_date)
        ).aggregate(total=Sum('bags'))['total'] or 0

        total_stock = inward_total + return_total - outward_total

        if key not in stock_dict:
            stock_dict[key] = {
                'Company': company.objects.get(id=combination['company']),
                'Goods': company_goods.objects.get(id=combination['company_goods']),
                'Company2': goods_company.objects.get(id=combination['goods_company']),
                'Stock': total_stock
            }

    context = {
        'data': list(stock_dict.values()),
        'selected_year': selected_year,
    }

    return render(request, 'transactions/list_stock.html', context)


@login_required(login_url='login')
def add_return(request):


    if request.method == 'POST':

        forms = supply_return_Form(request.POST)
        DC_date = request.POST.get('DC_date')

        forms = supply_return_Form(request.POST)

      

        if forms.is_valid():
            


            a = forms.cleaned_data['company']
            b = forms.cleaned_data['company_goods']
            c = forms.cleaned_data['goods_company']
            d = forms.cleaned_data['agent']
            e = forms.cleaned_data['bags']

            try:

                test = stock.objects.get(company = a, company_goods = b, goods_company = c)

                test.total_bag = test.total_bag + e
                test.save()
                forms.save()


                return JsonResponse({'status' : 'done'}, safe=False)


            except stock.DoesNotExist:

                test = stock.objects.create(company = a, company_goods = b, goods_company = c, total_bag = e)
                return JsonResponse({'status' : 'failed', 'error_msg' : 'stock does not exsist'}, safe=False)


        else:
            error = forms.errors.as_json()
            return JsonResponse({'error' : error}, safe=False)


    else:

        forms = supply_return_Form()

        context = {
            'form': forms
        }
        return render(request, 'transactions/add_return.html', context)



@login_required(login_url='login')
def list_return(request):

    year = request.GET.get('year')

    if year:

        date1 = year + '-04-01'
        date2 = str(int(year) + 1) + '-03-31'

        data = inward.objects.filter(DC_date__range=[date1, date2])
        data.extra(select={'DC_number':'SUBSTRING("DC_number",m,-)'}).order_by(Substr('DC_number',3))
    
    else:

        data = supply_return.objects.all()
        data.extra(select={'DC_number':'SUBSTRING("DC_number",m,-)'}).order_by(Substr('DC_number',3))

    supply_return_filter_data = supply_return_filter()

    # inward_filter_data = inward_filter()

    agent_name = request.GET.get('agent_name')

    if agent_name:

        data = data.filter(agent__name__icontains=agent_name)



    total_bags = data.aggregate(Sum('bags'))['bags__sum']


    page = request.GET.get('page', 1)
    paginator = Paginator(data, 50)

    
    try:
        data = paginator.page(page)
    except PageNotAnInteger:
        data = paginator.page(1)
    except EmptyPage:
        data = paginator.page(paginator.num_pages)



    company_data = company.objects.all()

    context = {
        'data': data,
        'company_data': company_data,
        'total_bags' : total_bags,
        'filter_return_supply' : supply_return_filter_data,
        'year' : year

    }

    return render(request, 'transactions/list_return.html', context)


@login_required(login_url='login')
def update_return(request, return_id):

    if request.method == 'POST':

        instance = supply_return.objects.get(id = return_id)

        forms = supply_return_Form(request.POST, instance=instance)

        if forms.is_valid():

            forms.save()


              
        else:

            instance = supply_return.objects.get(id = return_id)
            comapnyID = forms.instance.company.id
            comapny_goods_ID = forms.instance.company_goods.id
            goods_company_ID = forms.instance.goods_company.id
            agent_ID = forms.instance.agent.id

            context = {
                'form': forms,
                'comapnyID' : comapnyID,
                'comapny_goods_ID' : comapny_goods_ID,
                'goods_company_ID' : goods_company_ID,
                'agent_ID' : agent_ID       
            }
            return render(request, 'transactions/update_return.html', context)

    else:

        instance = supply_return.objects.get(id = return_id)
        forms = supply_return_Form(instance = instance)
        comapnyID = forms.instance.company.id
        comapny_goods_ID = forms.instance.company_goods.id
        goods_company_ID = forms.instance.goods_company.id
        agent_ID = forms.instance.agent.id

        context = {
            'form': forms,
            'comapnyID' : comapnyID,
            'comapny_goods_ID' : comapny_goods_ID,
            'goods_company_ID' : goods_company_ID,
            'agent_ID' : agent_ID       
        }
        return render(request, 'transactions/update_return.html', context)


@login_required(login_url='login')
def delete_return(request, return_id):

    try:
        supply_return.objects.get(id = return_id).delete()
        
        return HttpResponseRedirect(reverse('list_return_delete'))


    except:
        return HttpResponseRedirect(reverse('list_return_delete'))
 


def remove_timezone(dt):
    if isinstance(dt, datetime) and dt.tzinfo is not None:
        return dt.replace(tzinfo=None)
    return dt



# Function to handle encoding issues
def clean_string(s):
    if isinstance(s, str):
        return s.encode('ascii', 'ignore').decode('ascii')
    return s


from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os
from django.http import HttpResponse
from django.utils.timezone import localtime
from django.conf import settings
from datetime import datetime

@login_required(login_url='login')
def report_inward(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    import os

    counter = 1
    vals = []

    data = inward.objects.all().order_by("DC_number")
    filtered_data = inward_filter(request.GET, data).qs

    total_bags = filtered_data.aggregate(Sum('bags'))['bags__sum']

    filtered_data = list(filtered_data.values_list(
        'DC_number', 'agent__name', 'agent__place', 'agent__taluka',
        'agent__district', 'company_goods__name', 'goods_company__goods_company_name',
        'bags', 'DC_date__date', 'transport__name', 'LR_number', 'freight', 'id'
    ))

    filtered_data = list(map(list, filtered_data))

    # CSV Header
    headers = ["Serial", "DC Number", "Party Name", "Party Place", "Party Taluka", "Party District",
               "Crop", "Variety", "Packet", "Date", "Transport", "LR Number", "Freight", "id"]
    vals.append(headers)

    # CSV Rows
    for i in filtered_data:
        row = [
            counter, i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7],
            '%s/%s/%s' % (i[8].day, i[8].month, i[8].year), i[9], i[10], i[11], i[12]
        ]
        vals.append(row)
        counter += 1

    # Save XLSX
    wb = Workbook()
    ws = wb.active
    ws.title = "Inward Report"

    # Add headers to XLSX
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')

    # Add data rows
    for i, row in enumerate(filtered_data, start=1):
        ws.append([
            i, row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7],
            '%s/%s/%s' % (row[8].day, row[8].month, row[8].year), row[9], row[10], row[11], row[12]
        ])

    xlsx_name = "Report.xlsx"
    xlsx_path = os.path.join(BASE_DIR, 'static', 'xlsx', xlsx_name)
    wb.save(xlsx_path)

    vals_list = vals.copy()
    vals_list.pop(0)  # remove headers for display table

    context = {
        'data': vals_list,
        'total_bags': total_bags,
        'link': f"/static/xlsx/{xlsx_name}",
    }

    return render(request, 'report/inward_report.html', context)



@login_required(login_url='login')
def report_outward(request):
    counter = 1
    vals = []

    outward_data = outward.objects.all().order_by("DC_number")
    outward_filterd_data = outward_filter(request.GET, outward_data).qs

    total_bags = outward_filterd_data.aggregate(Sum('bags'))['bags__sum']

    outward_filterd_data = list(outward_filterd_data.values_list(
        'DC_number', 'agent__name', 'agent__place', 'agent__taluka', 'agent__district',
        'company_goods__name', 'goods_company__goods_company_name', 'bags', 'DC_date__date',
        'transport__name', 'LR_number', 'freight', 'id'
    ))

    outward_filterd_data = list(map(list, outward_filterd_data))

    # CSV Header
    headers = ["Serial", "DC Number", "Party Name", "Party Place", "Party Taluka", "Party District",
               "Crop", "Variety", "Packet", "Date", "Transport", "LR Number", "Freight", "id"]
    vals.append(headers)

    # CSV Rows
    for i in outward_filterd_data:
        row = [
            counter, i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7],
            '%s/%s/%s' % (i[8].day, i[8].month, i[8].year), i[9], i[10], i[11], i[12]
        ]
        vals.append(row)
        counter += 1

   

    # Save XLSX
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Outward Report"

    # Add headers to XLSX
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')

    # Add data rows
    for i, row in enumerate(outward_filterd_data, start=1):
        ws.append([
            i, row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7],
            '%s/%s/%s' % (row[8].day, row[8].month, row[8].year), row[9], row[10], row[11]
        ])

    xlsx_name = "Report.xlsx"
    xlsx_path = os.path.join(BASE_DIR, 'static', 'xlsx', xlsx_name)
    wb.save(xlsx_path)

    print(xlsx_path)

    vals_list = vals.copy()
    vals_list.pop(0)  # remove headers

    context = {
        'data': vals_list,
        'total_bags': total_bags,
        'link': xlsx_path,
    }

    return render(request, 'report/outward_report.html', context)


from django.db.models.functions import Substr

@login_required(login_url='login')
def report_supply_return(request):
    data = supply_return.objects.all()
    data.extra(select={'DC_number':'SUBSTRING("DC_number",m,-)'}).order_by(Substr('DC_number',3))


    filterd_data = supply_return_filter(request.GET, data)
    data = filterd_data.qs

    total_bags = data.aggregate(Sum('bags'))['bags__sum']

    filtered_data = list(data.values_list(
        'DC_number', 'agent__name', 'agent__place', 'agent__taluka', 'agent__district',
        'company_goods__name', 'goods_company__goods_company_name', 'bags', 'DC_date__date',
        'transport__name', 'LR_number', 'freight', 'id'
    ))

    vals = []
    headers = ["Serial", "DC Number", "Party Name", "Party Place", "Party Taluka", "Party District",
               "Crop", "Variety", "Packet", "Date", "Transport", "LR Number", "Freight", "id"]
    vals.append(headers)

    counteer = 1
    for i in filtered_data:
        row = [
            counteer, i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7],
            '%s/%s/%s' % (i[8].day, i[8].month, i[8].year), i[9], i[10], i[11], i[12]
        ]
        vals.append(row)
        counteer += 1

    # Save CSV
    csv_name = "Report.csv"
    csv_path = os.path.join(BASE_DIR, 'static', 'csv', csv_name)
    with open(csv_path, 'w', newline="") as f:
        writer = csv.writer(f)
        writer.writerows(vals)

    # Save XLSX
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Supply Return Report"

    # Add headers
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')

    # Add data rows
    for i, row in enumerate(filtered_data, start=1):
        ws.append([
            i, row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7],
            '%s/%s/%s' % (row[8].day, row[8].month, row[8].year), row[9], row[10], row[11]
        ])

    xlsx_name = "Report.xlsx"
    xlsx_path = os.path.join(BASE_DIR, 'static', 'xlsx', xlsx_name)
    wb.save(xlsx_path)

    vals_list = vals.copy()
    vals_list.pop(0)

    context = {
        'data': vals_list,
        'total_bags': total_bags,
        'link': csv_path,
        'link_xlsx': xlsx_path,
    }

    return render(request, 'report/supply_return_report.html', context)



@login_required(login_url='login')
def generate_report_stock(request):


    data = stock.objects.all()

    data_stock = stock_filter(request.GET, data)
    data_stock = data_stock.qs

    data1 = []
    data2 = []





    if data_stock:

        for i in data_stock:

         
            data1.append(i.company.company_name)
            data1.append(i.company_goods)
            data1.append(i.goods_company)
            data1.append(i.total_bag)




            data2.append(data1)

            data1 = []



    time =  str(datetime.now(ist))
    time = time.split('.')
    time = time[0].replace(':', '-')

    name = "Report.csv"
    path = os.path.join(BASE_DIR) + '\static\csv\\' + name
    with open(path,  'w', newline="") as f:
        writer = csv.writer(f)
        writer.writerows(data2)

    link = os.path.join(BASE_DIR) + '\static\csv\\' + name

    stock_filter_data = stock_filter()

    context = {
        'data': data2,
        'stock_filter' : stock_filter_data,
        'link' : link

    }

    return render(request, 'report/stock_report.html', context)




@login_required(login_url='login')
def generate_report_main(request):

    company_data = pd.DataFrame(list(company.objects.all().values('id', 'company_name')))
    company_data = dict(company_data.values)

    company_goods_data = pd.DataFrame(list(company_goods.objects.all().values('id', 'name')))
    company_goods_data = dict(company_goods_data.values)

    goods_company_data = pd.DataFrame(list(goods_company.objects.all().values('id', 'goods_company_name')))
    goods_company_data = dict(goods_company_data.values)

    agent_data = pd.DataFrame(list(agent.objects.all().values('id', 'name')))
    agent_data = dict(agent_data.values)
    
    agent_data2 = pd.DataFrame(list(agent.objects.all().values('id', 'name', 'place', 'taluka', 'district')))

    # return_data = supply_return.objects.all()

    outward_data = outward.objects.all()
    supply_return_data = supply_return.objects.all()
    outward_filterd_data = outward_filter(request.GET, outward_data)
    supply_return_filterd_data = outward_filter(request.GET, supply_return_data)

    if outward_data and not supply_return_data:
        
        # outward sum
        df2 = pd.DataFrame(list(outward_filterd_data.qs.values()))
        sum__ = df2.groupby(['company_id', 'company_goods_id', 'goods_company_id', 'agent_id']).sum().reset_index()
       

        sum__['bags_x'] = sum__['bags']
        sum__['bags_y'] = None
        sum__['bags_z'] = sum__['bags']
        sum__['company_id'] = sum__['company_id'].map(company_data)
        sum__['company_goods_id'] = sum__['company_goods_id'].map(company_goods_data)
        sum__['goods_company_id'] = sum__['goods_company_id'].map(goods_company_data)
        sum__['agent_id'] = sum__['agent_id'].map(agent_data)
        

        out = (sum__.merge(agent_data2, left_on='agent_id', right_on='name').reindex(columns=['agent_id', 'place', 'taluka', 'district', 'company_id', 'company_goods_id', 'goods_company_id', 'bags_x', 'bags_y', 'bags_z']))



    elif supply_return_data and not outward_data:
        #return sum
        df3 = pd.DataFrame(list(supply_return_filterd_data.qs.values()))

        sum__2 = df3.groupby(['company_id', 'company_goods_id', 'goods_company_id', 'agent_id']).sum().reset_index()
     
        sum__2['bags_x'] = None
        sum__2['bags_y'] = sum__2['bags']
        sum__2['bags_z'] = sum__2['bags']
        sum__2['company_id'] = sum__2['company_id'].map(company_data)
        sum__2['company_goods_id'] = sum__2['company_goods_id'].map(company_goods_data)
        sum__2['goods_company_id'] = sum__2['goods_company_id'].map(goods_company_data)
        sum__2['agent_id'] = sum__2['agent_id'].map(agent_data)
        

        out = (sum__2.merge(agent_data2, left_on='agent_id', right_on='name').reindex(columns=['agent_id', 'place', 'taluka', 'district', 'company_id', 'company_goods_id', 'goods_company_id', 'bags_x', 'bags_y', 'bags_z']))




    else:

        df2 = pd.DataFrame(list(outward_filterd_data.qs.values()))

        df3 = pd.DataFrame(list(supply_return_filterd_data.qs.values()))


        sum__ = df2.groupby(['company_id', 'company_goods_id', 'goods_company_id', 'agent_id']).sum().reset_index()
        sum__2 = df3.groupby(['company_id', 'company_goods_id', 'goods_company_id', 'agent_id']).sum().reset_index()


        final_ou = pd.merge(sum__, sum__2, on=['company_id', 'company_goods_id', 'goods_company_id', 'agent_id'], how="outer")[['company_id', 'company_goods_id', 'goods_company_id', 'agent_id', 'bags_x', 'bags_y']]
        final_ou['bags_z'] = final_ou.fillna(0)['bags_x'] - final_ou.fillna(0)['bags_y']

    
        final_ou['company_id'] = final_ou['company_id'].map(company_data)
        final_ou['company_goods_id'] = final_ou['company_goods_id'].map(company_goods_data)
        final_ou['goods_company_id'] = final_ou['goods_company_id'].map(goods_company_data)
        final_ou['agent_id'] = final_ou['agent_id'].map(agent_data)

        # print(final_ou)
        # here
        out = (final_ou.merge(agent_data2, left_on='agent_id', right_on='name').reindex(columns=['company_id', 'agent_id', 'place', 'taluka', 'district', 'company_goods_id', 'goods_company_id', 'bags_x', 'bags_y', 'bags_z']))

       

    vals = out.values

    time =  str(datetime.now(ist))
    time = time.split('.')
    time = time[0].replace(':', '-')
        


    
    vals_list = (vals.tolist())
    vals1 = []
    vals1.append("Party Name")
    vals1.append("Party Place")
    vals1.append("Party Taluka")
    vals1.append("Party District")
    vals1.append("Company")
    vals1.append("Crop")
    vals1.append("Variety")
    vals1.append('Supply Packet')
    vals1.append('Return Packet')
    vals1.append('Net Packet')

    vals_list.insert(0, vals1)

    time =  str(datetime.now(ist))
    time = time.split('.')
    time = time[0].replace(':', '-')


    name = "Report.csv"
    path = os.path.join(BASE_DIR) + '\static\csv\\' + name
    with open(path,  'w', newline="") as f:
        writer = csv.writer(f)
        writer.writerows(vals_list)


    link = os.path.join(BASE_DIR) + '\static\csv\\' + name


    outward_filter_data = outward_filter()

    vals_list = (vals.tolist())

    company_data = company.objects.all()

    context = {
        'data' : vals_list,
        'filter_outward' : outward_filter_data,
        'link' : link,
        'company_data' : company_data

    }

    return render(request, 'report/main_report.html', context)
    

@login_required(login_url='login')
def generate_report_daily(request):
    pd.set_option('display.float_format', '{:.2f}'.format)

    #DC_date_start__date=2022-02-28&DC_date_end__date=2022-02-28

    company_data = pd.DataFrame(list(company.objects.all().values('id', 'company_name')))
    company_data = dict(company_data.values)

    company_goods_data = pd.DataFrame(list(company_goods.objects.all().values('id', 'name')))
    company_goods_data = dict(company_goods_data.values)

    goods_company_data = pd.DataFrame(list(goods_company.objects.all().values('id', 'goods_company_name')))
    goods_company_data = dict(goods_company_data.values)

    inward_data = inward.objects.filter(DC_date__date = date.today())
    outward_data = outward.objects.filter(DC_date__date = date.today())
    supply_return_data = supply_return.objects.filter(DC_date__date = date.today())
    inward_filterd_data = inward_filter(request.GET, inward_data)
    outward_data_filterd_data = outward_filter(request.GET, outward_data)
    supply_return_filterd_data = outward_filter(request.GET, supply_return_data)
    if inward_filterd_data.qs:
        # inward sum
        df = pd.DataFrame(list(inward_filterd_data.qs.values()))
        sum__ = df.groupby(['company_id', 'company_goods_id', 'goods_company_id']).sum().reset_index()
    else:
        sum__ = pd.DataFrame(columns=['company_id', 'company_goods_id', 'goods_company_id', 'id', 'agent_id', 'bags', 'DC_number'])

    if outward_data_filterd_data.qs:
        # outward sum
        df2 = pd.DataFrame(list(outward_data_filterd_data.qs.values()))
        sum__2 = df2.groupby(['company_id', 'company_goods_id', 'goods_company_id']).sum().reset_index()
        
    else:
        sum__2 = pd.DataFrame(columns=['company_id', 'company_goods_id', 'goods_company_id', 'id', 'agent_id', 'bags', 'DC_number'])

    if supply_return_filterd_data.qs:
        
        #return sum
        df3 = pd.DataFrame(list(supply_return_filterd_data.qs.values()))
        sum__3 = df3.groupby(['company_id', 'company_goods_id', 'goods_company_id']).sum().reset_index()

    else:
        sum__3 = pd.DataFrame(columns=['company_id', 'company_goods_id', 'goods_company_id', 'id', 'agent_id', 'bags', 'DC_number'])

    #stock
    sum__4 = pd.DataFrame(list(stock.objects.all().values()))
   

    data_frames = [sum__, sum__2, sum__3, sum__4]

    ada = reduce(lambda  left,right: pd.merge(left,right,on=['company_id', 'company_goods_id', 'goods_company_id'], how='outer'), data_frames)[['company_id', 'company_goods_id', 'goods_company_id', 'bags_x', 'bags_y', 'bags', 'total_bag']]
  
   

    # m = pd.merge(sum__, sum__2, on=['company_id', 'company_goods_id', 'goods_company_id'], how="outer")[['company_id', 'company_goods_id', 'goods_company_id',  'bags_x', 'bags_y']]
    
    # print('m')
    # print(m)
    # m1 = pd.merge(m, sum__3, on=['company_id', 'company_goods_id', 'goods_company_id'], how="outer")[['company_id', 'company_goods_id', 'goods_company_id',  'bags_x', 'bags_y', 'bags']]
   
    # print('m1')
    # print(m1)
    # m2 = pd.merge(m1, sum__4, on=['company_id', 'company_goods_id', 'goods_company_id'], how="outer")[['company_id', 'company_goods_id', 'goods_company_id',  'bags_x', 'bags_y', 'bags', 'total_bag']]

    # print('m2')
    # print(m2)

    final_out = ada

    final_out['company_id'] = final_out['company_id'].map(company_data)
    final_out['company_goods_id'] = final_out['company_goods_id'].map(company_goods_data)
    final_out['goods_company_id'] = final_out['goods_company_id'].map(goods_company_data)

    vals = final_out.values
    
    time =  str(datetime.now(ist))
    time = time.split('.')
    time = time[0].replace(':', '-')

    name = "Report.csv"
    path = os.path.join(BASE_DIR) + '\static\csv\\' + name
    with open(path,  'w', newline="") as f:
        writer = csv.writer(f)
        writer.writerows(vals)

    outward_filter_data = outward_filter()

    link = os.path.join(BASE_DIR) + '\static\csv\\' + name

    vals_list = (vals.tolist())


    context = {
        'data': vals_list,
        'filter_outward' : outward_filter_data,
        'link' : link

    }

    return render(request, 'report/daily_report.html', context)



@login_required(login_url='login')
def generate_report_monthly(request):

    pd.set_option('display.float_format', '{:.2f}'.format)
    print('--------------soukdgsbvhgdvei')

    #DC_date_start__date=2022-02-28&DC_date_end__date=2022-02-28

    company_data = pd.DataFrame(list(company.objects.all().values('id', 'company_name')))
    company_data = dict(company_data.values)

    company_goods_data = pd.DataFrame(list(company_goods.objects.all().values('id', 'name')))
    company_goods_data = dict(company_goods_data.values)

    goods_company_data = pd.DataFrame(list(goods_company.objects.all().values('id', 'goods_company_name')))
    goods_company_data = dict(goods_company_data.values)

    inward_data = inward.objects.all()
    outward_data = outward.objects.all()
    supply_return_data = supply_return.objects.all()
    inward_filterd_data = inward_filter(request.GET, inward_data)
    outward_data_filterd_data = outward_filter(request.GET, outward_data)
    supply_return_filterd_data = outward_filter(request.GET, supply_return_data)

    if inward_filterd_data.qs:
        # inward sum
        df = pd.DataFrame(list(inward_filterd_data.qs.values()))
        sum__ = df.groupby(['company_id', 'company_goods_id', 'goods_company_id']).sum().reset_index()
    else:
        sum__ = pd.DataFrame(columns=['company_id', 'company_goods_id', 'goods_company_id', 'id', 'agent_id', 'bags', 'DC_number'])

    if outward_data_filterd_data.qs:
        # outward sum
        df2 = pd.DataFrame(list(outward_data_filterd_data.qs.values()))
        sum__2 = df2.groupby(['company_id', 'company_goods_id', 'goods_company_id']).sum().reset_index()
        
    else:
        sum__2 = pd.DataFrame(columns=['company_id', 'company_goods_id', 'goods_company_id', 'id', 'agent_id', 'bags', 'DC_number'])

    if supply_return_filterd_data.qs:
        
        #return sum
        df3 = pd.DataFrame(list(supply_return_filterd_data.qs.values()))
        sum__3 = df3.groupby(['company_id', 'company_goods_id', 'goods_company_id']).sum().reset_index()

    else:
        sum__3 = pd.DataFrame(columns=['company_id', 'company_goods_id', 'goods_company_id', 'id', 'agent_id', 'bags', 'DC_number'])

    #stock
    sum__4 = pd.DataFrame(list(stock.objects.all().values()))
   

    data_frames = [sum__, sum__2, sum__3, sum__4]

    ada = reduce(lambda  left,right: pd.merge(left,right,on=['company_id', 'company_goods_id', 'goods_company_id'], how='outer'), data_frames)[['company_id', 'company_goods_id', 'goods_company_id', 'bags_x', 'bags_y', 'bags', 'total_bag']]


    # m = pd.merge(sum__, sum__2, on=['company_id', 'company_goods_id', 'goods_company_id'], how="outer")[['company_id', 'company_goods_id', 'goods_company_id',  'bags_x', 'bags_y']]
    
    # print('m')
    # print(m)
    # m1 = pd.merge(m, sum__3, on=['company_id', 'company_goods_id', 'goods_company_id'], how="outer")[['company_id', 'company_goods_id', 'goods_company_id',  'bags_x', 'bags_y', 'bags']]
   
    # print('m1')
    # print(m1)
    # m2 = pd.merge(m1, sum__4, on=['company_id', 'company_goods_id', 'goods_company_id'], how="outer")[['company_id', 'company_goods_id', 'goods_company_id',  'bags_x', 'bags_y', 'bags', 'total_bag']]

    # print('m2')
    # print(m2)

    final_out = ada

    final_out['company_id'] = final_out['company_id'].map(company_data)
    final_out['company_goods_id'] = final_out['company_goods_id'].map(company_goods_data)
    final_out['goods_company_id'] = final_out['goods_company_id'].map(goods_company_data)

    vals = final_out.values
    
    time =  str(datetime.now(ist))
    time = time.split('.')
    time = time[0].replace(':', '-')

    name = "Report.csv"
    path = os.path.join(BASE_DIR) + '\static\csv\\' + name
    with open(path,  'w', newline="") as f:
        writer = csv.writer(f)
        writer.writerows(vals)

    outward_filter_data = outward_filter()

    link = os.path.join(BASE_DIR) + '\static\csv\\' + name

    vals_list = (vals.tolist())


    context = {
        'data': vals_list,
        'filter_outward' : outward_filter_data,
        'link' : link

    }

    return render(request, 'report/customize.html', context)



@login_required(login_url='login')
def download(request):
    # fill these variables with real values


    file_path = os.path.join(BASE_DIR, 'static', 'xlsx', 'Report.xlsx')
    if os.path.exists(file_path):
        return FileResponse(open(file_path, 'rb'), as_attachment=True, filename='Report.xlsx')
    else:
        return HttpResponse("File not found", status=404)


@login_required(login_url='login')
def delete_dashboard(request):

    return render(request, 'delete/dashbaord.html')


# delete view


def list_inward_delete(request):

    year = request.GET.get('year')

    if year:

        date1 = year + '-04-01'
        date2 = str(int(year) + 1) + '-03-31'

        data = inward.objects.filter(DC_date__range=[date1, date2]).order_by("DC_number")
    
    else:

        data = inward.objects.filter().order_by("DC_number")

    total_bags = data.aggregate(Sum('bags'))['bags__sum']
    

    inward_filter_data = inward_filter()

    company_data = company.objects.all()

    page = request.GET.get('page', 1)
    paginator = Paginator(data, 50)

    try:
        data = paginator.page(page)
    except PageNotAnInteger:
        data = paginator.page(1)
    except EmptyPage:
        data = paginator.page(paginator.num_pages)

    context = {
        'data': data,
        'company_data' : company_data,
        'total_bags' : total_bags,
        'filter_inward' : inward_filter_data,
        'year' : year
    }

    return render(request, 'delete/list_inward_delete.html', context)


def list_outward_delete(request):

    
    year = request.GET.get('year')

    if year:

        date1 =year + '-04-01'
        date2 = str(int(year) + 1) + '-03-31'

        data = outward.objects.filter(DC_date__range=[date1, date2]).order_by("DC_number")
    
    else:

        data = outward.objects.filter().order_by('DC_number')


    agent_name = request.GET.get('agent_name')

    if agent_name:

        data = data.filter(agent__name__icontains=agent_name)

    total_bags = data.aggregate(Sum('bags'))['bags__sum']
    

    outward_filter_data = outward_filter()

    company_data = company.objects.all()



    page = request.GET.get('page', 1)
    paginator = Paginator(data, 50)

    try:
        data = paginator.page(page)
    except PageNotAnInteger:
        data = paginator.page(1)
    except EmptyPage:
        data = paginator.page(paginator.num_pages)



    context = {
        'data': data,
        'filter_outward' : outward_filter_data,
        'total_bags' : total_bags,
        'company_data' : company_data,
        'year' : year

    }

    return render(request, 'delete/list_outward_delete.html', context)

def list_return_delete(request):

   
    year = request.GET.get('year')

    if year:

        date1 = year + '-04-01'
        date2 = str(int(year) + 1) + '-03-31'

        data = inward.objects.filter(DC_date__range=[date1, date2])
        data.extra(select={'DC_number':'SUBSTRING("DC_number",m,-)'}).order_by(Substr('DC_number',3))
    
    else:

        data = supply_return.objects.all()
        data.extra(select={'DC_number':'SUBSTRING("DC_number",m,-)'}).order_by(Substr('DC_number',3))

    supply_return_filter_data = supply_return_filter()

    # inward_filter_data = inward_filter()

    agent_name = request.GET.get('agent_name')

    if agent_name:

        data = data.filter(agent__name__icontains=agent_name)



    total_bags = data.aggregate(Sum('bags'))['bags__sum']


    page = request.GET.get('page', 1)
    paginator = Paginator(data, 50)

    
    try:
        data = paginator.page(page)
    except PageNotAnInteger:
        data = paginator.page(1)
    except EmptyPage:
        data = paginator.page(paginator.num_pages)



    company_data = company.objects.all()

    context = {
        'data': data,
        'company_data': company_data,
        'total_bags' : total_bags,
        'filter_return_supply' : supply_return_filter_data,
        'year' : year

    }



    return render(request, 'delete/list_return_delete.html', context)






